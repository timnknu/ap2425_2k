# Програма аналізує текст із doc-файлу, виявляє усі використані в ньому текстові символи, залишає лише українські літери, 
# рахує їхню кількість та частотність (у відсотках), а результати записує як талицю у xlsx-файл

import docx
import openpyxl
import re

from openpyxl.styles import PatternFill

d = docx.Document("sample-text.docx")
all_text = ''
for p in d.paragraphs:
    for r in p.runs:
        all_text += r.text.lower() + ' '
all_text = re.sub(' {1,}', ' ', all_text)
all_symbols = set(all_text)
print('all_symbols:', sorted(all_symbols))
all_ukr_symbols = ['а', 'б', 'в', 'г', 'д', 'е', 'ж', 'з', 'и', 'й', 'к', 'л', 'м', 'н', 'о', 'п', 'р', 'с', 'т', 'у', 'ф', 'х', 'ц', 'ч', 'ш', 'щ', 'ь', 'ю', 'я', 'є', 'і', 'ї', 'ґ']
all_symbols = all_symbols.intersection(set(all_ukr_symbols))
count = {}
for c in all_symbols:
    count[c] = all_text.count(c)
tot_ukr_symbols = sum(count.values())
# for c in sorted(:
#     count[c] = all_text.count(c)
# version1:
# def get_val(tpl):
#     return -tpl[1]
# print(sorted(count.items(), key=get_val))
# version2:
#print(sorted(count.items(), key=lambda tpl: -tpl[1]))
for k,v in sorted(count.items(), key=lambda tpl: -tpl[1]):
    print(k, v, v/tot_ukr_symbols*100)
wb = openpyxl.Workbook()
ws = wb.active
sorted_freqs = sorted(count.items(), key=lambda tpl: -tpl[1])
max_freq = sorted_freqs[0][1]/tot_ukr_symbols*100 # всі кортежі мають вигляд (символ, частота) ; беремо частоту із першого кортежу; рахуємо її як відсоток від загальної кількості українських символів
for rw,(k,v) in enumerate(sorted_freqs):
    obj = ws.cell(rw+1,1)
    obj.value = k
    obj = ws.cell(rw+1,2)
    obj.value = v
    obj = ws.cell(rw+1,3)
    freq = v/tot_ukr_symbols*100
    obj.value = freq

    # Створюємо код кольору, який залежить від частоти символу
    color_hex = 'FFFF' + f'{int(255 - 255 * freq/max_freq):02x}'
    # для цього ми перетворюємо частоту у "насиченість" кольору; коли частота символу максимальна (freq/max_freq==1),
    # то колір буде мати код FFFF00 (жовтий), а коли мінімальна (freq/max_freq==0) - то колір буде мати код FFFFFF (білий)
    # Запис кольору у форматі RGB: FFFFFF - білий, 000000 - чорний, FF0000 - червоний, 00FF00 - зелений, 0000FF - синій,
    # FFFF00 - жовтий, 00FFFF - бірюзовий, FF00FF - пурпурний тощо
    # print(color_hex)
    obj.fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid") # використовуємо властивість fill об'єкта-клітинки, щоб змінити колір фону клітинки

wb.save('freq.xlsx')
